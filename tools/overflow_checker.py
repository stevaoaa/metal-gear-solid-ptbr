#!/usr/bin/env python3
"""
Script para analisar overflows em traduções.
Calcula tamanhos em bytes, identifica overflows e destaca células problemáticas.
SEMPRE recalcula as colunas de tamanho para garantir sincronização.

Uso:
    python overflow_checker.py --radio
    python overflow_checker.py --stage
    python overflow_checker.py --vox
    python overflow_checker.py --zmovie
    python overflow_checker.py --all
    python overflow_checker.py arquivo.csv
"""

import pandas as pd
import numpy as np
import os
import sys
import argparse
from datetime import datetime

# Adiciona o diretório pai ao sys.path para permitir importações relativas
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from util.logger_config import setup_logger


# Inicializa o logger personalizado
logger = setup_logger()

# Mapeamento dos parâmetros para arquivos
ARQUIVOS_TRADUZIDOS = {
    'demo': './translated/strings_DEMO_traduzido.csv',
    'radio': './translated/strings_RADIO_traduzido.csv',
    'stage': './translated/strings_STAGE_traduzido.csv', 
    'vox': './translated/strings_VOX_traduzido.csv',
    'zmovie': './translated/strings_ZMOVIE_traduzido.csv'
}

def calcular_tamanho_bytes(texto):
    """Calcula o tamanho em bytes usando codificação latin-1"""
    if pd.isna(texto) or texto == '':
        return 0
    return len(str(texto).encode('latin-1', errors='replace'))

def analisar_overflows(arquivo_csv, arquivo_saida=None, nome_dataset=None):
    """
    Analisa overflows em arquivo CSV de traduções.
    SEMPRE recalcula as colunas de tamanho para garantir sincronização.
    
    Args:
        arquivo_csv (str): Caminho para o arquivo CSV
        arquivo_saida (str): Caminho para arquivo de saída (opcional)
        nome_dataset (str): Nome do dataset para logs (opcional)
    
    Returns:
        tuple: (DataFrame processado, estatísticas de overflow)
    """
    
    dataset_info = f" ({nome_dataset.upper()})" if nome_dataset else ""
    logger.info(f" ANALISADOR DE OVERFLOWS{dataset_info}")
    logger.info(f" Arquivo: {arquivo_csv}")
    logger.info("=" * 60)
    
    # Verifica se arquivo existe
    if not os.path.exists(arquivo_csv):
        logger.error(f" Arquivo não encontrado: {arquivo_csv}")
        return None, None
    
    # Carrega dados
    try:
        df = pd.read_csv(arquivo_csv, delimiter='\t')
        logger.info(f" Arquivo carregado: {len(df)} linhas")
    except Exception as e:
        logger.error(f" Erro ao carregar arquivo: {e}")
        return None, None
    
    # Verifica colunas necessárias
    colunas_necessarias = ['texto_traduzido']
    colunas_faltando = [col for col in colunas_necessarias if col not in df.columns]
    
    if colunas_faltando:
        logger.warning(f" Colunas obrigatórias não encontradas: {colunas_faltando}")
        return None, None
    
    # SEMPRE recalcula tamanho_bytes baseado na coluna 'texto'
    if 'texto' in df.columns:
        logger.info(" Recalculando tamanho_bytes baseado na coluna 'texto'...")
        df['tamanho_bytes'] = df['texto'].apply(calcular_tamanho_bytes)
        coluna_referencia = 'tamanho_bytes'
    else:
        logger.warning(" Coluna 'texto' não encontrada - análise sem referência")
        coluna_referencia = None
    
    # SEMPRE recalcula tamanho_bytes_traduzido baseado na coluna 'texto_traduzido'
    logger.info(" Recalculando tamanho_bytes_traduzido baseado na coluna 'texto_traduzido'...")
    df['tamanho_bytes_traduzido'] = df['texto_traduzido'].apply(calcular_tamanho_bytes)
    
    # Remove colunas antigas que podem estar inconsistentes
    colunas_antigas = ['tamanho_bytes_traducao', 'tamanho_bytes_original', 'bytes_disponiveis']
    for col in colunas_antigas:
        if col in df.columns:
            logger.info(f" Removendo coluna antiga inconsistente: '{col}'")
            df = df.drop(columns=[col])
    
    # Identifica overflows
    if coluna_referencia:
        logger.info(f" Identificando overflows comparando 'tamanho_bytes_traduzido' com '{coluna_referencia}'...")
        
        # Calcula diferenças
        df['diferenca_bytes'] = df['tamanho_bytes_traduzido'] - df[coluna_referencia]
        df['tem_overflow'] = df['diferenca_bytes'] > 0
        df['percentual_diferenca'] = ((df['diferenca_bytes'] / df[coluna_referencia]) * 100).round(1)
        
        # Substitui inf por 0 quando referência é 0
        df['percentual_diferenca'] = df['percentual_diferenca'].replace([np.inf, -np.inf], 0)
        
        # Filtra apenas linhas com tradução
        df_com_traducao = df[df['texto_traduzido'].notna() & (df['texto_traduzido'] != '')]
        
        # Estatísticas
        total_traducoes = len(df_com_traducao)
        overflows = df_com_traducao['tem_overflow'].sum()
        
        if total_traducoes > 0:
            taxa_overflow = (overflows / total_traducoes) * 100
            
            logger.info(f"\n ESTATÍSTICAS DE OVERFLOW:")
            logger.info(f"   Total de traduções: {total_traducoes}")
            logger.info(f"   Overflows detectados: {overflows}")
            logger.info(f"   Taxa de overflow: {taxa_overflow:.1f}%")
            
            # Estatísticas de tamanho
            logger.info(f"\n ESTATÍSTICAS DE TAMANHO:")
            logger.info(f"   Texto original - Média: {df_com_traducao['tamanho_bytes'].mean():.1f} bytes")
            logger.info(f"   Texto original - Máximo: {df_com_traducao['tamanho_bytes'].max()} bytes")
            logger.info(f"   Tradução - Média: {df_com_traducao['tamanho_bytes_traduzido'].mean():.1f} bytes")
            logger.info(f"   Tradução - Máximo: {df_com_traducao['tamanho_bytes_traduzido'].max()} bytes")
            
            if overflows > 0:
                maior_overflow = df_com_traducao['diferenca_bytes'].max()
                media_overflow = df_com_traducao[df_com_traducao['tem_overflow']]['diferenca_bytes'].mean()
                
                logger.info(f"\n DETALHES DOS OVERFLOWS:")
                logger.info(f"   Maior overflow: +{maior_overflow} bytes")
                logger.info(f"   Overflow médio: +{media_overflow:.1f} bytes")
                
                # Top 10 overflows
                logger.info(f"\n TOP 10 OVERFLOWS:")
                top_overflows = df_com_traducao[df_com_traducao['tem_overflow']].nlargest(10, 'diferenca_bytes')
                
                for idx, row in top_overflows.iterrows():
                    texto_orig = str(row.get('texto', ''))[:40] + ('...' if len(str(row.get('texto', ''))) > 40 else '')
                    texto_trad = str(row['texto_traduzido'])[:40] + ('...' if len(str(row['texto_traduzido'])) > 40 else '')
                    
                    logger.info(f"   {idx:4d}: +{row['diferenca_bytes']:2d}b ({row['percentual_diferenca']:+.1f}%) | Orig:{row['tamanho_bytes']}b → Trad:{row['tamanho_bytes_traduzido']}b")
                    logger.info(f"         Original: '{texto_orig}'")
                    logger.info(f"         Tradução: '{texto_trad}'")
                    logger.info("")
        
        stats = {
            'dataset': nome_dataset or 'desconhecido',
            'total_traducoes': total_traducoes,
            'overflows': overflows,
            'taxa_overflow': taxa_overflow if total_traducoes > 0 else 0,
            'maior_overflow': df_com_traducao['diferenca_bytes'].max() if overflows > 0 else 0,
            'media_overflow': df_com_traducao[df_com_traducao['tem_overflow']]['diferenca_bytes'].mean() if overflows > 0 else 0
        }
    else:
        # Apenas estatísticas básicas sem comparação
        df_com_traducao = df[df['texto_traduzido'].notna() & (df['texto_traduzido'] != '')]
        total_traducoes = len(df_com_traducao)
        
        logger.info(f"\n ESTATÍSTICAS BÁSICAS:")
        logger.info(f"   Total de traduções: {total_traducoes}")
        logger.info(f"   Tamanho médio: {df_com_traducao['tamanho_bytes_traduzido'].mean():.1f} bytes")
        logger.info(f"   Tamanho máximo: {df_com_traducao['tamanho_bytes_traduzido'].max()} bytes")
        logger.info(f"   Tamanho mínimo: {df_com_traducao['tamanho_bytes_traduzido'].min()} bytes")
        
        stats = {
            'dataset': nome_dataset or 'desconhecido',
            'total_traducoes': total_traducoes,
            'overflows': 0,
            'taxa_overflow': 0,
            'maior_overflow': 0,
            'media_overflow': 0
        }
    
    # Define arquivo de saída
    if arquivo_saida is None:
        nome_base = os.path.splitext(arquivo_csv)[0]
        arquivo_saida = f"{nome_base}_analisado.csv"
    
    # Salva arquivo processado
    logger.info(f"\n Salvando arquivo processado...")
    try:
        df.to_csv(arquivo_saida, index=False, sep='\t', encoding='utf-8')
        logger.info(f" Arquivo salvo: {arquivo_saida}")
        logger.info(f" Colunas de tamanho sincronizadas:")
        logger.info(f"   - tamanho_bytes: calculado de 'texto'")
        logger.info(f"   - tamanho_bytes_traduzido: calculado de 'texto_traduzido'")
    except Exception as e:
        logger.error(f" Erro ao salvar: {e}")
    
    return df, stats

def gerar_relatorio_excel(df, arquivo_excel, coluna_referencia=None, nome_dataset=None):
    """
    Gera relatório Excel com formatação e destaque de overflows.
    
    Args:
        df (DataFrame): Dados processados
        arquivo_excel (str): Caminho para arquivo Excel de saída
        coluna_referencia (str): Nome da coluna de referência para comparação
        nome_dataset (str): Nome do dataset
    """
    
    try:
        # Importa bibliotecas necessárias para Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.info(" Para gerar Excel com formatação, instale: pip install openpyxl")
            return False
        
        logger.info(f" Gerando relatório Excel...")
        
        # Cria workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Overflow {nome_dataset.upper()}" if nome_dataset else "Análise de Overflows"
        
        # Define estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        overflow_fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")  # Vermelho claro
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Amarelo claro
        ok_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")        # Verde claro
        
        # Adiciona dados
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Formata cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Identifica colunas importantes
        colunas = {col: idx + 1 for idx, col in enumerate(df.columns)}
        
        col_traducao = colunas.get('texto_traduzido')
        col_tamanho_trad = colunas.get('tamanho_bytes_traduzido')
        col_tem_overflow = colunas.get('tem_overflow')
        col_diferenca = colunas.get('diferenca_bytes')
        
        # Aplica formatação condicional
        if col_tem_overflow and col_diferenca:
            logger.info(" Aplicando formatação condicional...")
            
            for row_idx in range(2, len(df) + 2):  # Pula cabeçalho
                tem_overflow = ws.cell(row=row_idx, column=col_tem_overflow).value
                diferenca = ws.cell(row=row_idx, column=col_diferenca).value
                
                if tem_overflow:
                    # Overflow crítico (>5 bytes) - Vermelho
                    if diferenca and diferenca > 5:
                        fill_color = overflow_fill
                    # Overflow moderado (1-5 bytes) - Amarelo
                    else:
                        fill_color = warning_fill
                    
                    # Destaca células da tradução e tamanho
                    if col_traducao:
                        ws.cell(row=row_idx, column=col_traducao).fill = fill_color
                    if col_tamanho_trad:
                        ws.cell(row=row_idx, column=col_tamanho_trad).fill = fill_color
                    
                    # Destaca diferença
                    if col_diferenca:
                        ws.cell(row=row_idx, column=col_diferenca).fill = fill_color
                        ws.cell(row=row_idx, column=col_diferenca).font = Font(bold=True)
                else:
                    # Sem overflow - Verde claro
                    if col_diferenca and diferenca is not None and diferenca <= 0:
                        ws.cell(row=row_idx, column=col_diferenca).fill = ok_fill
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Define largura com limites
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adiciona aba de estatísticas
        ws_stats = wb.create_sheet("Estatísticas")
        
        # Cabeçalho das estatísticas
        stats_data = [
            ["Métrica", "Valor"],
            ["Dataset", nome_dataset.upper() if nome_dataset else "N/A"],
            ["Total de traduções", len(df[df['texto_traduzido'].notna() & (df['texto_traduzido'] != '')])],
        ]
        
        if 'tem_overflow' in df.columns:
            overflows = df['tem_overflow'].sum()
            taxa = (overflows / len(df)) * 100 if len(df) > 0 else 0
            
            stats_data.extend([
                ["Overflows detectados", overflows],
                ["Taxa de overflow (%)", f"{taxa:.1f}%"],
            ])
            
            if overflows > 0:
                stats_data.extend([
                    ["Maior overflow (bytes)", df['diferenca_bytes'].max()],
                    ["Overflow médio (bytes)", f"{df[df['tem_overflow']]['diferenca_bytes'].mean():.1f}"],
                ])
        
        # Adiciona informações sobre sincronização
        stats_data.extend([
            ["", ""],
            ["SINCRONIZAÇÃO DE DADOS", ""],
            ["tamanho_bytes", "Recalculado de 'texto'"],
            ["tamanho_bytes_traduzido", "Recalculado de 'texto_traduzido'"],
            ["", ""],
            ["LEGENDA DE CORES", ""],
            ["Overflow crítico (>5 bytes)", "Vermelho"],
            ["Overflow moderado (1-5 bytes)", "Amarelo"], 
            ["Sem overflow", "Verde"],
        ])
        
        for row in stats_data:
            ws_stats.append(row)
        
        # Formata aba de estatísticas
        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Adiciona cores na legenda
        ws_stats['B13'].fill = overflow_fill  # Vermelho
        ws_stats['B14'].fill = warning_fill   # Amarelo
        ws_stats['B15'].fill = ok_fill       # Verde
        
        # Salva arquivo
        wb.save(arquivo_excel)
        logger.info(f" Relatório Excel salvo: {arquivo_excel}")
        
        return True
        
    except Exception as e:
        logger.error(f" Erro ao gerar Excel: {e}")
        return False

def processar_arquivo(dataset, gerar_excel=False):
    """
    Processa um arquivo específico do dataset.
    
    Args:
        dataset (str): Nome do dataset (radio, stage, vox, etc.)
        gerar_excel (bool): Se deve gerar relatório Excel
    
    Returns:
        dict: Estatísticas do processamento
    """
    
    arquivo_csv = ARQUIVOS_TRADUZIDOS.get(dataset.lower())
    
    if not arquivo_csv:
        logger.error(f" Dataset '{dataset}' não reconhecido!")
        logger.info(f" Datasets disponíveis: {', '.join(ARQUIVOS_TRADUZIDOS.keys())}")
        return None
    
    if not os.path.exists(arquivo_csv):
        logger.error(f" Arquivo não encontrado: {arquivo_csv}")
        return None
    
    # Executa análise
    df, stats = analisar_overflows(arquivo_csv, nome_dataset=dataset)
    
    if df is not None and stats is not None:
        # Gera Excel se solicitado
        if gerar_excel:
            nome_base = os.path.splitext(arquivo_csv)[0]
            arquivo_excel = f"{nome_base}_relatorio.xlsx"
            
            coluna_ref = 'tamanho_bytes' if 'tamanho_bytes' in df.columns else None
            gerar_relatorio_excel(df, arquivo_excel, coluna_ref, dataset)
        
        return stats
    
    return None

def processar_todos_arquivos(gerar_excel=False):
    """
    Processa todos os arquivos de tradução disponíveis.
    
    Args:
        gerar_excel (bool): Se deve gerar relatórios Excel
    
    Returns:
        dict: Resumo de todos os processamentos
    """
    
    logger.info(" PROCESSANDO TODOS OS DATASETS")
    logger.info("=" * 60)
    
    resultados = {}
    total_traducoes = 0
    total_overflows = 0
    
    for dataset in ARQUIVOS_TRADUZIDOS.keys():
        arquivo = ARQUIVOS_TRADUZIDOS[dataset]
        
        if os.path.exists(arquivo):
            logger.info(f"\n Processando {dataset.upper()}...")
            stats = processar_arquivo(dataset, gerar_excel)
            
            if stats:
                resultados[dataset] = stats
                total_traducoes += stats['total_traducoes']
                total_overflows += stats['overflows']
        else:
            logger.warning(f" Arquivo {dataset.upper()} não encontrado: {arquivo}")
    
    # Resumo geral
    logger.info(f"\n" + "=" * 60)
    logger.info(f" RESUMO GERAL")
    logger.info("=" * 60)
    
    if resultados:
        taxa_geral = (total_overflows / total_traducoes * 100) if total_traducoes > 0 else 0
        
        logger.info(f"   Total de datasets processados: {len(resultados)}")
        logger.info(f"   Total de traduções: {total_traducoes}")
        logger.info(f"   Total de overflows: {total_overflows}")
        logger.info(f"   Taxa geral de overflow: {taxa_geral:.1f}%")
        logger.info("")
        
        # Tabela por dataset
        logger.info("   DETALHES POR DATASET:")
        logger.info("   " + "-" * 50)
        for dataset, stats in resultados.items():
            logger.info(f"   {dataset.upper():8s}: {stats['overflows']:3d}/{stats['total_traducoes']:5d} overflows ({stats['taxa_overflow']:5.1f}%)")
        
        return {
            'datasets_processados': len(resultados),
            'total_traducoes': total_traducoes,
            'total_overflows': total_overflows,
            'taxa_geral': taxa_geral,
            'detalhes': resultados
        }
    else:
        logger.error("   Nenhum arquivo foi processado!")
        return None

def main():
    """Função principal com parsing de argumentos"""
    
    parser = argparse.ArgumentParser(
        description='Analisador de Overflows em Traduções',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Exemplos de uso:
  {sys.argv[0]} --radio                    # Analisa strings_RADIO_traduzido.csv
  {sys.argv[0]} --stage --excel            # Analisa STAGE e gera Excel
  {sys.argv[0]} --all                      # Analisa todos os datasets
  {sys.argv[0]} --all --excel              # Analisa todos e gera Excel
  {sys.argv[0]} --file arquivo.csv         # Analisa arquivo específico

Datasets disponíveis: {', '.join(ARQUIVOS_TRADUZIDOS.keys())}
        """
    )
    
    # Argumentos para datasets específicos
    parser.add_argument('--demo', action='store_true', help='Analisa strings_DEMO_traduzido.csv')
    parser.add_argument('--radio', action='store_true', help='Analisa strings_RADIO_traduzido.csv')
    parser.add_argument('--stage', action='store_true', help='Analisa strings_STAGE_traduzido.csv')
    parser.add_argument('--vox', action='store_true', help='Analisa strings_VOX_traduzido.csv')
    parser.add_argument('--zmovie', action='store_true', help='Analisa strings_ZMOVIE_traduzido.csv')
    parser.add_argument('--all', action='store_true', help='Analisa todos os datasets disponíveis')
    
    # Arquivo específico (agora como argumento nomeado)
    parser.add_argument('--file', '-f', help='Arquivo CSV específico para analisar')
    
    # Opções adicionais
    parser.add_argument('--excel', action='store_true', help='Gera relatório Excel com formatação')
    parser.add_argument('-o', '--output', help='Arquivo de saída (apenas para arquivo específico)')
    
    args = parser.parse_args()
    
    # Identifica qual dataset foi selecionado
    datasets_selecionados = []
    if args.demo: datasets_selecionados.append('demo')
    if args.radio: datasets_selecionados.append('radio')
    if args.stage: datasets_selecionados.append('stage')
    if args.vox: datasets_selecionados.append('vox')
    if args.zmovie: datasets_selecionados.append('zmovie')
    
    # Verifica se múltiplos datasets foram selecionados
    if len(datasets_selecionados) > 1:
        logger.error(" Erro: Apenas um dataset pode ser selecionado por vez")
        logger.info(f" Datasets selecionados: {', '.join(datasets_selecionados)}")
        return
    
    # Se nenhum argumento foi fornecido, mostra ajuda
    if not any([args.demo, args.radio, args.stage, args.vox, args.zmovie, args.all, args.file]):
        parser.print_help()
        return
    
    # Processa baseado nos argumentos
    if args.all:
        # Processa todos os arquivos
        logger.info(" Processando todos os datasets...")
        processar_todos_arquivos(args.excel)
        
    elif datasets_selecionados:
        # Dataset específico foi selecionado
        dataset = datasets_selecionados[0]
        logger.info(f" Processando dataset: {dataset.upper()}")
        
        stats = processar_arquivo(dataset, args.excel)
        
        if stats:
            logger.info(f"\n Análise de {dataset.upper()} concluída!")
            
            if stats['overflows'] > 0:
                logger.info(f" {stats['overflows']} overflows detectados - verifique as células destacadas")
            else:
                logger.info(f" Nenhum overflow detectado!")
        
    elif args.file:
        # Arquivo específico fornecido
        if not os.path.exists(args.file):
            logger.error(f" Arquivo não encontrado: {args.file}")
            return
        
        logger.info(f" Processando arquivo específico: {args.file}")
        df, stats = analisar_overflows(args.file, args.output)
        
        if df is not None and stats is not None:
            if args.excel:
                nome_base = os.path.splitext(args.file)[0]
                arquivo_excel = f"{nome_base}_relatorio.xlsx"
                
                coluna_ref = 'tamanho_bytes' if 'tamanho_bytes' in df.columns else None
                gerar_relatorio_excel(df, arquivo_excel, coluna_ref)
            
            logger.info(f"\n Análise concluída!")
            
            if stats['overflows'] > 0:
                logger.info(f" {stats['overflows']} overflows detectados")
            else:
                logger.info(f" Nenhum overflow detectado!")

if __name__ == "__main__":
    """
    # Datasets específicos
    python .\tools\overflow_checker.py --vox
    python .\tools\overflow_checker.py --radio
    python .\tools\overflow_checker.py --stage --excel

    # Todos os datasets
    python .\tools\overflow_checker.py --all
    python .\tools\overflow_checker.py --all --excel

    # Arquivo específico
    python .\tools\overflow_checker.py --file meuarquivo.csv
    python .\tools\overflow_checker.py -f meuarquivo.csv --excel
    """
    
    main()